# utils.py - 通用工具函数与UI组件
import tkinter as tk
from tkinter import ttk
import uuid
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from data_manager import DataManager
from openpyxl.utils import get_column_letter


# ============ ScrollableFrame 组件（来自components.py） ============
class ScrollableFrame(ttk.Frame):
    """自定义滚动框架"""

    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, bg="white")
        self.scrollbar_y = ttk.Scrollbar(
            self, orient="vertical", command=self.canvas.yview
        )
        self.scrollbar_x = ttk.Scrollbar(
            self, orient="horizontal", command=self.canvas.xview
        )
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )

        # 创建 window 并保存 id，以便在 canvas 大小改变时同步内部 frame 宽度
        self._canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # 使内部 frame 宽度随 canvas 宽度变化（避免内容被压缩为很窄）
        def _on_canvas_configure(event):
            try:
                self.canvas.itemconfig(self._canvas_window, width=event.width)
            except Exception:
                pass

        self.canvas.bind("<Configure>", _on_canvas_configure)
        self.canvas.configure(
            yscrollcommand=self.scrollbar_y.set,
            xscrollcommand=self.scrollbar_x.set,
        )

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar_y.pack(side="right", fill="y")
        self.scrollbar_x.pack(side="bottom", fill="x")


# ============ 渲染函数 ============
class UIRenderer:
    """负责各步骤的数据表渲染"""

    @staticmethod
    def render_env_frame(step3_frame, env_data, selected_models, model_test_type_map, vendor_str, main_scroll):
        """渲染步骤3的测试环境框架 - 改进的布局"""
        for w in step3_frame.winfo_children():
            w.destroy()

        # 创建可滚动的内部框架
        scroll_frm = ScrollableFrame(step3_frame)
        scroll_frm.pack(fill=tk.BOTH, expand=True)
        content_frm = scroll_frm.scrollable_frame

        # 如果没有 env_data，则用一个占位的动态条目渲染一次，方便用户手动填写
        render_list = env_data if env_data else [
            {
                "model": "",
                "test_type": "",
                "vendor": "",
                "gpu": "",
                "gpu_count": "",
                "dataset": "",
                "tool": "",
                "is_dynamic": True,
                "id": str(uuid.uuid4()),
            }
        ]

        # 如果 env_data 为空，先创建一条默认条目并加入（确保 validate 能看到数据）
        if not env_data:
            env_data.append({
                "model": "",
                "test_type": "",
                "vendor": "",
                "gpu": "",
                "gpu_count": "",
                "dataset": "",
                "tool": "",
                "is_dynamic": True,
                "id": str(uuid.uuid4()),
            })

        # helper: 安全写入 env_data
        def _write_env_value(idx, key, value):
            while idx >= len(env_data):
                env_data.append({
                    "model": "",
                    "test_type": "",
                    "vendor": "",
                    "gpu": "",
                    "gpu_count": "",
                    "dataset": "",
                    "tool": "",
                    "is_dynamic": True,
                    "id": str(uuid.uuid4()),
                })
            env_data[idx][key] = value

        # 为每一行环境配置创建一个分组框
        for row_idx, data in enumerate(env_data):
            # 为每个环境创建一个卡片式框架
            card_frm = ttk.LabelFrame(
                content_frm,
                text=f"配置 #{row_idx + 1}: {data.get('model', '')} - {data.get('test_type', '')}",
                padding=10
            )
            card_frm.pack(fill=tk.BOTH, expand=True, padx=5, pady=8)

            # 使用 grid 布局：使单行控件横向扩展，避免堆叠
            card_frm.columnconfigure(1, weight=1)
            card_frm.columnconfigure(3, weight=1)
            card_frm.columnconfigure(5, weight=1)

            # 第一行：模型、测试类型
            if data.get("is_dynamic"):
                ttk.Label(card_frm, text="模型：").grid(row=0, column=0, sticky="w", padx=5, pady=2)
                model_combo = ttk.Combobox(card_frm, values=selected_models, state="readonly")
                model_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
                model_combo.set(data["model"])

                ttk.Label(card_frm, text="测试类型：").grid(row=0, column=2, sticky="w", padx=5, pady=2)
                tt_combo = ttk.Combobox(card_frm, state="readonly")
                tt_combo.grid(row=0, column=3, sticky="ew", padx=5, pady=2)
                tt_combo.set(data["test_type"])

                def on_model_change(e, combo=tt_combo, combo_model=model_combo, idx=row_idx):
                    combo.config(values=model_test_type_map.get(combo_model.get(), []))
                    _write_env_value(idx, "model", combo_model.get())
                    _write_env_value(idx, "test_type", combo.get())

                model_combo.bind("<<ComboboxSelected>>", on_model_change)
            else:
                ttk.Label(card_frm, text="模型：").grid(row=0, column=0, sticky="w", padx=5, pady=2)
                ttk.Label(card_frm, text=data["model"]).grid(row=0, column=1, sticky="w", padx=5, pady=2)

                ttk.Label(card_frm, text="测试类型：").grid(row=0, column=2, sticky="w", padx=5, pady=2)
                ttk.Label(card_frm, text=data["test_type"]).grid(row=0, column=3, sticky="w", padx=5, pady=2)

            # 第二行：厂家、GPU型号
            vendor_list = UIRenderer._parse_vendor_str(vendor_str)
            vendor_names = [v[0] for v in vendor_list]
            vendor_gpus = {v[0]: v[1] for v in vendor_list}

            if data.get("is_dynamic"):
                ttk.Label(card_frm, text="厂家：").grid(row=1, column=0, sticky="w", padx=5, pady=2)
                vendor_combo = ttk.Combobox(card_frm, values=vendor_names, state="readonly")
                vendor_combo.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
                vendor_combo.set(data.get("vendor", ""))

                ttk.Label(card_frm, text="GPU型号：").grid(row=1, column=2, sticky="w", padx=5, pady=2)
                gpu_lbl = ttk.Label(card_frm, text=data.get("gpu", ""))
                gpu_lbl.grid(row=1, column=3, sticky="w", padx=5, pady=2)

                def on_vendor_change(e, lbl=gpu_lbl, vendor_combo_ref=vendor_combo, idx=row_idx):
                    lbl.config(text=vendor_gpus.get(vendor_combo_ref.get(), ""))
                    _write_env_value(idx, "vendor", vendor_combo_ref.get())
                    _write_env_value(idx, "gpu", vendor_gpus.get(vendor_combo_ref.get(), ""))

                vendor_combo.bind("<<ComboboxSelected>>", on_vendor_change)
            else:
                ttk.Label(card_frm, text="厂家：").grid(row=1, column=0, sticky="w", padx=5, pady=2)
                ttk.Label(card_frm, text=data.get("vendor", "")).grid(row=1, column=1, sticky="w", padx=5, pady=2)

                ttk.Label(card_frm, text="GPU型号：").grid(row=1, column=2, sticky="w", padx=5, pady=2)
                ttk.Label(card_frm, text=data.get("gpu", "")).grid(row=1, column=3, sticky="w", padx=5, pady=2)

            # 第三行：GPU数量、数据集、测试工具
            ttk.Label(card_frm, text="GPU数量：").grid(row=2, column=0, sticky="w", padx=5, pady=2)
            gpu_var = tk.StringVar(value=str(data.get("gpu_count", "1")))
            gpu_cnt = tk.Spinbox(card_frm, from_=1, to=100, textvariable=gpu_var, width=6)
            gpu_cnt.grid(row=2, column=1, sticky="w", padx=5, pady=2)

            # 使用变量追踪，确保通过键入或上下箭头都能更新 env_data
            def _on_gpu_var_change(*args, idx=row_idx, var=gpu_var):
                _write_env_value(idx, "gpu_count", var.get())

            try:
                gpu_var.trace_add("write", _on_gpu_var_change)
            except AttributeError:
                # Python <3.6 fallback
                gpu_var.trace("w", _on_gpu_var_change)

            ttk.Label(card_frm, text="数据集：").grid(row=2, column=2, sticky="w", padx=5, pady=2)
            dataset_var = tk.StringVar(value=data.get("dataset", ""))
            dataset = ttk.Entry(card_frm, textvariable=dataset_var)
            dataset.grid(row=2, column=3, sticky="ew", padx=5, pady=2)

            def _on_dataset_var(*a, idx=row_idx, var=dataset_var):
                _write_env_value(idx, "dataset", var.get())

            try:
                dataset_var.trace_add("write", _on_dataset_var)
            except AttributeError:
                dataset_var.trace("w", _on_dataset_var)

            ttk.Label(card_frm, text="测试工具：").grid(row=2, column=4, sticky="w", padx=5, pady=2)
            tool_var = tk.StringVar(value=data.get("tool", ""))
            tool = ttk.Entry(card_frm, textvariable=tool_var)
            tool.grid(row=2, column=5, sticky="ew", padx=5, pady=2)

            def _on_tool_var(*a, idx=row_idx, var=tool_var):
                _write_env_value(idx, "tool", var.get())

            try:
                tool_var.trace_add("write", _on_tool_var)
            except AttributeError:
                tool_var.trace("w", _on_tool_var)

            # 删除按钮放在第四行，靠左；仅当是真实 env_data 时显示
            if env_data:
                del_btn = ttk.Button(
                    card_frm,
                    text="🗑️ 删除此配置",
                    command=lambda idx=row_idx: UIRenderer._del_env_row(idx, env_data, step3_frame, selected_models, model_test_type_map, vendor_str, main_scroll)
                )
                del_btn.grid(row=3, column=0, sticky="w", padx=5, pady=6)

        # 新增按钮
        add_btn_frm = ttk.Frame(content_frm)
        add_btn_frm.pack(fill=tk.X, padx=5, pady=10)
        ttk.Button(
            add_btn_frm,
            text="➕ 新增环境配置",
            command=lambda: UIRenderer._add_env_row(env_data, selected_models, model_test_type_map, vendor_str, step3_frame, main_scroll),
        ).pack(side=tk.LEFT, padx=5)

        # 刷新滚动
        main_scroll.canvas.update_idletasks()
        main_scroll.canvas.configure(scrollregion=main_scroll.canvas.bbox("all"))

    @staticmethod
    def render_models_and_test_types(
        model_type_frame, app_ref
    ):
        """渲染模型输入和测试类型选择"""
        from config import PERF_FIELDS_MAP
        
        # 清空旧控件
        for w in model_type_frame.winfo_children():
            w.destroy()
        
        # 获取测试类型列表
        yaml_path = "model_config.yaml"
        model_names, test_types = DataManager.load_models(yaml_path, app_ref)
        
        # 表头
        ttk.Label(model_type_frame, text="模型名称", font=("", 9, "bold")).grid(
            row=0, column=0, padx=10, pady=5, sticky="w"
        )
        ttk.Label(model_type_frame, text="测试类型", font=("", 9, "bold")).grid(
            row=0, column=1, padx=10, pady=5, sticky="w"
        )
        ttk.Label(model_type_frame, text="操作", font=("", 9, "bold")).grid(
            row=0, column=2, padx=10, pady=5, sticky="w"
        )
        
        # 初始化模型数据（如果未初始化）
        if not app_ref.model_input_data:
            app_ref.model_input_data = [
                {
                    "model_name": tk.StringVar(value=""),
                    "test_types": {tt: tk.IntVar() for tt in test_types},
                    "id": str(__import__('uuid').uuid4())
                }
            ]
        
        # 渲染每一行模型输入
        for row_idx, model_data in enumerate(app_ref.model_input_data, 1):
            # 模型名称输入框
            model_entry = ttk.Entry(model_type_frame, width=20, textvariable=model_data["model_name"])
            model_entry.grid(row=row_idx, column=0, padx=10, pady=5)
            
            # 测试类型容器
            tt_container = ttk.Frame(model_type_frame)
            tt_container.grid(row=row_idx, column=1, padx=10, pady=5, sticky="w")
            
            # 横向排列测试类型复选框
            col_idx = 0
            for tt in test_types:
                var = model_data["test_types"][tt]
                chk = ttk.Checkbutton(tt_container, text=tt, variable=var)
                chk.grid(row=0, column=col_idx, padx=5, pady=1, sticky="w")
                col_idx += 1
            
            # 操作按钮
            btn_frame = ttk.Frame(model_type_frame)
            btn_frame.grid(row=row_idx, column=2, padx=10, pady=5)
            
            ttk.Button(
                btn_frame,
                text="➕ 新增",
                command=lambda: UIRenderer._add_model_row(model_data, app_ref, test_types, model_type_frame),
                width=8
            ).pack(side=tk.LEFT, padx=2)
            
            if len(app_ref.model_input_data) > 1:
                ttk.Button(
                    btn_frame,
                    text="删除",
                    command=lambda mid=model_data["id"]: UIRenderer._del_model_row(mid, app_ref, test_types, model_type_frame),
                    width=6
                ).pack(side=tk.LEFT, padx=2)
        
        model_type_frame.columnconfigure(0, weight=0)
        model_type_frame.columnconfigure(1, weight=1)
        model_type_frame.columnconfigure(2, weight=0)
    
    @staticmethod
    def _add_model_row(model_data, app_ref, test_types, model_type_frame):
        """新增模型行"""
        new_model_data = {
            "model_name": tk.StringVar(value=""),
            "test_types": {tt: tk.IntVar() for tt in test_types},
            "id": str(__import__('uuid').uuid4())
        }
        current_idx = app_ref.model_input_data.index(model_data)
        app_ref.model_input_data.insert(current_idx + 1, new_model_data)
        UIRenderer.render_models_and_test_types(model_type_frame, app_ref)
        model_type_frame.after(100, lambda: model_type_frame.yview_moveto(1.0))
    
    @staticmethod
    def _del_model_row(model_id, app_ref, test_types, model_type_frame):
        """删除模型行"""
        app_ref.model_input_data = [
            m for m in app_ref.model_input_data if m["id"] != model_id
        ]
        if not app_ref.model_input_data:
            app_ref.model_input_data = [
                {
                    "model_name": tk.StringVar(value=""),
                    "test_types": {tt: tk.IntVar() for tt in test_types},
                    "id": str(__import__('uuid').uuid4())
                }
            ]
        UIRenderer.render_models_and_test_types(model_type_frame, app_ref)

    @staticmethod
    def render_pk_frame(pk_frame, pk_data, app_ref):
        """渲染PK指标框架"""
        for w in pk_frame.winfo_children():
            w.destroy()

        headers = ["序号", "模型", "测试类型", "PK指标", "操作"]
        for i, h in enumerate(headers):
            ttk.Label(pk_frame, text=h, font=("", 9, "bold")).grid(
                row=0, column=i, padx=4, pady=5
            )

        for row_idx, pk_row in enumerate(pk_data, 1):
            ttk.Label(pk_frame, text=str(row_idx - 1)).grid(
                row=row_idx, column=0, padx=4, pady=3
            )
            ttk.Label(pk_frame, text=pk_row["model"]).grid(
                row=row_idx, column=1, padx=4, pady=3
            )
            ttk.Label(pk_frame, text=pk_row["test_type"]).grid(
                row=row_idx, column=2, padx=4, pady=3
            )

            pk_combo = ttk.Combobox(pk_frame, width=60, state="normal")
            pk_combo.config(values=pk_row["pk_options"])
            pk_combo.grid(row=row_idx, column=3, padx=4, pady=3, sticky="ew")
            pk_combo.set(pk_row["selected_pk"])

            def on_pk_change(e, row_id=pk_row["id"], combo=pk_combo):
                for r in pk_data:
                    if r["id"] == row_id:
                        r["selected_pk"] = combo.get()
                        break

            pk_combo.bind("<<ComboboxSelected>>", on_pk_change)

            btn_frame = ttk.Frame(pk_frame)
            btn_frame.grid(row=row_idx, column=4, padx=4, pady=3)

            ttk.Button(
                btn_frame,
                text="➕ 新增行",
                command=lambda row_data=pk_row: UIRenderer._add_pk_row(row_data, pk_data, pk_frame, app_ref),
                width=8,
            ).pack(side=tk.LEFT, padx=2)

            ttk.Button(
                btn_frame,
                text="删除",
                command=lambda row_id=pk_row["id"]: UIRenderer._del_pk_row(row_id, pk_data, pk_frame, app_ref),
                width=6,
            ).pack(side=tk.LEFT, padx=2)

        app_ref.main_scroll.canvas.update_idletasks()
        app_ref.main_scroll.canvas.configure(
            scrollregion=app_ref.main_scroll.canvas.bbox("all")
        )

    @staticmethod
    def render_perf_frame(perf_frame, perf_data, app_ref):
        """渲染性能数据框架"""
        from config import PERF_FIELDS_MAP

        for w in perf_frame.winfo_children():
            w.destroy()

        test_type_groups = {}
        for perf_row in perf_data:
            tt = perf_row["test_type"]
            if tt not in test_type_groups:
                test_type_groups[tt] = []
            test_type_groups[tt].append(perf_row)

        block_row = 0
        for test_type, rows in test_type_groups.items():
            tt_frm = ttk.LabelFrame(perf_frame, text=f"📊 {test_type}")
            tt_frm.grid(row=block_row, column=0, sticky="nsew", padx=5, pady=8, columnspan=4)
            block_row += 1

            input_fields, calc_fields = PERF_FIELDS_MAP.get(test_type, ([], []))

            headers = ["序号", "模型", "厂家", "数据集"] + input_fields + calc_fields + ["操作"]
            for col_idx, h in enumerate(headers):
                lbl = ttk.Label(tt_frm, text=h, font=("", 9, "bold"))
                lbl.grid(row=0, column=col_idx, padx=3, pady=5, sticky="nsew")

            for row_idx, perf_row in enumerate(rows, 1):
                col_idx = 0

                ttk.Label(tt_frm, text=str(row_idx - 1)).grid(
                    row=row_idx, column=col_idx, padx=3, pady=3
                )
                col_idx += 1

                ttk.Label(tt_frm, text=perf_row["model"]).grid(
                    row=row_idx, column=col_idx, padx=3, pady=3
                )
                col_idx += 1

                ttk.Label(tt_frm, text=perf_row["vendor"]).grid(
                    row=row_idx, column=col_idx, padx=3, pady=3
                )
                col_idx += 1

                ttk.Label(tt_frm, text=perf_row["dataset"], relief="sunken", width=18).grid(
                    row=row_idx, column=col_idx, padx=3, pady=3
                )
                col_idx += 1

                for field in input_fields:
                    entry = ttk.Entry(tt_frm, width=18)
                    entry.grid(row=row_idx, column=col_idx, padx=3, pady=3)
                    entry.insert(0, perf_row["input_values"][field])

                    def on_input_change(e, row_id=perf_row["id"], field=field, entry=entry):
                        for r in perf_data:
                            if r["id"] == row_id:
                                r["input_values"][field] = entry.get()
                                break

                    entry.bind("<KeyRelease>", on_input_change)
                    col_idx += 1

                for field in calc_fields:
                    lbl = ttk.Label(
                        tt_frm, text=perf_row["calc_values"][field], relief="sunken", width=18
                    )
                    lbl.grid(row=row_idx, column=col_idx, padx=3, pady=3)
                    col_idx += 1

                btn_frame = ttk.Frame(tt_frm)
                btn_frame.grid(row=row_idx, column=col_idx, padx=2, pady=3, sticky="w")

                ttk.Button(
                    btn_frame,
                    text="➕ 新增行",
                    command=lambda row_data=perf_row: UIRenderer._add_perf_row(row_data, perf_data, perf_frame, app_ref),
                ).pack(side=tk.LEFT, padx=2)

                ttk.Button(
                    btn_frame,
                    text="删除",
                    command=lambda row_id=perf_row["id"]: UIRenderer._del_perf_row(row_id, perf_data, perf_frame, app_ref),
                ).pack(side=tk.LEFT, padx=2)

        calc_btn = ttk.Button(
            perf_frame,
            text="📊 计算文本/图文推理吞吐数据",
            command=app_ref._calculate_throughput,
        )
        calc_btn.grid(row=block_row, column=0, padx=5, pady=10, sticky="w", columnspan=4)

        app_ref.main_scroll.canvas.update_idletasks()
        app_ref.main_scroll.canvas.configure(
            scrollregion=app_ref.main_scroll.canvas.bbox("all")
        )

    @staticmethod
    def render_problem_frame(problem_frame, problem_data, app_ref):
        """渲染项目问题框架"""
        for w in problem_frame.winfo_children():
            w.destroy()

        headers = ["序号", "问题分类", "问题描述", "责任人", "解决方案", "操作"]
        for col_idx, h in enumerate(headers):
            lbl = ttk.Label(problem_frame, text=h, font=("", 9, "bold"))
            lbl.grid(row=0, column=col_idx, padx=8, pady=5, sticky="nsew")

        for row_idx, problem_row in enumerate(problem_data, 1):
            col_idx = 0

            ttk.Label(problem_frame, text=str(row_idx - 1)).grid(
                row=row_idx, column=col_idx, padx=8, pady=3
            )
            col_idx += 1

            category_combo = ttk.Combobox(
                problem_frame,
                values=["项目问题", "技术问题"],
                state="readonly",
                width=12,
            )
            category_combo.grid(row=row_idx, column=col_idx, padx=8, pady=3)
            category_combo.set(problem_row["category"])

            def on_category_change(e, row_id=problem_row["id"], combo=category_combo):
                for r in problem_data:
                    if r["id"] == row_id:
                        r["category"] = combo.get()
                        break

            category_combo.bind("<<ComboboxSelected>>", on_category_change)
            col_idx += 1

            desc_entry = ttk.Entry(problem_frame, width=40)
            desc_entry.grid(row=row_idx, column=col_idx, padx=8, pady=3)
            desc_entry.insert(0, problem_row["description"])

            def on_desc_change(e, row_id=problem_row["id"], entry=desc_entry):
                for r in problem_data:
                    if r["id"] == row_id:
                        r["description"] = entry.get()
                        break

            desc_entry.bind("<KeyRelease>", on_desc_change)
            col_idx += 1

            person_entry = ttk.Entry(problem_frame, width=15)
            person_entry.grid(row=row_idx, column=col_idx, padx=8, pady=3)
            person_entry.insert(0, problem_row["person"])

            def on_person_change(e, row_id=problem_row["id"], entry=person_entry):
                for r in problem_data:
                    if r["id"] == row_id:
                        r["person"] = entry.get()
                        break

            person_entry.bind("<KeyRelease>", on_person_change)
            col_idx += 1

            solution_entry = ttk.Entry(problem_frame, width=40)
            solution_entry.grid(row=row_idx, column=col_idx, padx=8, pady=3)
            solution_entry.insert(0, problem_row["solution"])

            def on_solution_change(e, row_id=problem_row["id"], entry=solution_entry):
                for r in problem_data:
                    if r["id"] == row_id:
                        r["solution"] = entry.get()
                        break

            solution_entry.bind("<KeyRelease>", on_solution_change)
            col_idx += 1

            btn_frame = ttk.Frame(problem_frame)
            btn_frame.grid(row=row_idx, column=col_idx, padx=8, pady=3)

            ttk.Button(
                btn_frame,
                text="➕ 新增行",
                command=lambda row_data=problem_row: UIRenderer._add_problem_row(row_data, problem_data, problem_frame, app_ref),
                width=8,
            ).pack(side=tk.LEFT, padx=2)

            ttk.Button(
                btn_frame,
                text="删除",
                command=lambda row_id=problem_row["id"]: UIRenderer._del_problem_row(row_id, problem_data, problem_frame, app_ref),
                width=6,
            ).pack(side=tk.LEFT, padx=2)

        problem_frame.columnconfigure(2, weight=2)
        problem_frame.columnconfigure(4, weight=2)

        app_ref.main_scroll.canvas.update_idletasks()
        app_ref.main_scroll.canvas.configure(
            scrollregion=app_ref.main_scroll.canvas.bbox("all")
        )

    # ========== 辅助方法 ==========
    @staticmethod
    def _parse_vendor_str(vendor_str):
        """解析厂家字符串"""
        vendors = []
        if not vendor_str.strip():
            return vendors
        for item in vendor_str.split("、"):
            item = item.strip()
            if "（" in item and "）" in item:
                name = item.split("（")[0].strip()
                gpu = item.split("（")[1].replace("）", "").strip()
                if name and gpu:
                    vendors.append((name, gpu))
        return vendors

    @staticmethod
    def _add_env_row(env_data, selected_models, model_test_type_map, vendor_str, step3_frame, main_scroll):
        """新增环境行"""
        from data_manager import DataManager
        vendor_list = DataManager.parse_vendor_str(vendor_str)
        env_data.append({
            "model": selected_models[0] if selected_models else "",
            "test_type": "",
            "vendor": vendor_list[0][0] if vendor_list else "",
            "gpu": vendor_list[0][1] if vendor_list else "",
            "gpu_count": "",
            "dataset": "",
            "tool": "",
            "is_dynamic": True,
            "id": str(uuid.uuid4()),
        })
        UIRenderer.render_env_frame(step3_frame, env_data, selected_models, model_test_type_map, vendor_str, main_scroll)
        main_scroll.canvas.yview_moveto(1.0)

    @staticmethod
    def _del_env_row(idx, env_data, step3_frame, selected_models, model_test_type_map, vendor_str, main_scroll):
        """删除环境行"""
        if 0 <= idx < len(env_data):
            del env_data[idx]
            UIRenderer.render_env_frame(step3_frame, env_data, selected_models, model_test_type_map, vendor_str, main_scroll)

    @staticmethod
    def _add_pk_row(row_data, pk_data, pk_frame, app_ref):
        """新增PK指标行"""
        new_row = {
            "id": str(uuid.uuid4()),
            "model": row_data["model"],
            "test_type": row_data["test_type"],
            "pk_options": row_data["pk_options"],
            "selected_pk": "",
        }
        current_idx = pk_data.index(row_data)
        pk_data.insert(current_idx + 1, new_row)
        UIRenderer.render_pk_frame(pk_frame, pk_data, app_ref)

    @staticmethod
    def _del_pk_row(row_id, pk_data, pk_frame, app_ref):
        """删除PK指标行"""
        for i, r in enumerate(pk_data):
            if r["id"] == row_id:
                del pk_data[i]
                break
        UIRenderer.render_pk_frame(pk_frame, pk_data, app_ref)

    @staticmethod
    def _add_perf_row(row_data, perf_data, perf_frame, app_ref):
        """新增性能数据行"""
        new_row = {
            "id": str(uuid.uuid4()),
            "model": row_data["model"],
            "test_type": row_data["test_type"],
            "vendor": row_data["vendor"],
            "dataset": row_data["dataset"],
            "gpu_count": row_data["gpu_count"],
            "input_fields": row_data["input_fields"],
            "calc_fields": row_data["calc_fields"],
            "input_values": {f: "" for f in row_data["input_fields"]},
            "calc_values": {f: "" for f in row_data["calc_fields"]},
        }
        current_idx = perf_data.index(row_data)
        perf_data.insert(current_idx + 1, new_row)
        UIRenderer.render_perf_frame(perf_frame, perf_data, app_ref)

    @staticmethod
    def _del_perf_row(row_id, perf_data, perf_frame, app_ref):
        """删除性能数据行"""
        for i, r in enumerate(perf_data):
            if r["id"] == row_id:
                del perf_data[i]
                break
        UIRenderer.render_perf_frame(perf_frame, perf_data, app_ref)

    @staticmethod
    def _add_problem_row(row_data, problem_data, problem_frame, app_ref):
        """新增问题行"""
        new_row = {
            "id": str(uuid.uuid4()),
            "category": "",
            "description": "",
            "person": "",
            "solution": "",
        }
        current_idx = problem_data.index(row_data)
        problem_data.insert(current_idx + 1, new_row)
        UIRenderer.render_problem_frame(problem_frame, problem_data, app_ref)

    @staticmethod
    def _del_problem_row(row_id, problem_data, problem_frame, app_ref):
        """删除问题行"""
        for i, r in enumerate(problem_data):
            if r["id"] == row_id:
                del problem_data[i]
                break
        if not problem_data:
            problem_data.append({
                "id": str(uuid.uuid4()),
                "category": "",
                "description": "",
                "person": "",
                "solution": "",
            })
        UIRenderer.render_problem_frame(problem_frame, problem_data, app_ref)


# ============ 样式工具函数 ============
def load_yaml_config(yaml_path):
    """读取YAML配置，返回模型列表和测试类型列表"""
    if not os.path.exists(yaml_path):
        # 创建默认配置
        create_default_yaml(yaml_path)
    
    try:
        with open(yaml_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        
        model_list = config.get("model_names", [])
        test_type_list = config.get("test_types", [])
        
        return model_list, test_type_list
    except yaml.YAMLError as e:
        messagebox.showerror("YAML格式错误", f"语法错误：{str(e)}\n请用2个空格缩进")
        return [], []
    except Exception as e:
        messagebox.showerror("读取失败", f"配置文件错误：{str(e)}")
        return [], []


def create_default_yaml(yaml_path):
    """创建默认YAML配置"""
    config = {
        "model_names": ["DeepSeek-R1", "yolov11", "qwen14B"],
        "test_types": ["文本推理", "图文推理", "图像识别", "预训练", "lora微调", "全参微调"]
    }
    try:
        with open(yaml_path, "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False)
    except Exception as e:
        messagebox.showerror("创建失败", f"无法创建配置文件：{str(e)}")


def get_excel_styles():
    """返回Excel样式配置"""
    return {
        "header_font": Font(bold=True, size=11),
        "header_fill": PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
        "border": Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True)
    }


def validate_env_data(env_entries):
    """验证环境数据是否完整"""
    for idx, env in enumerate(env_entries, 1):
        required_fields = {
            "GPU数量": env.get("gpu_count"),
            "数据集": env.get("dataset"),
            "测试工具": env.get("tool")
        }
        for field_name, field_widget in required_fields.items():
            if field_widget and hasattr(field_widget, 'get'):
                if not field_widget.get().strip():
                    messagebox.showerror("填写不完整", f"第{idx}行 {field_name} 为必填项")
                    return False
    return True


def parse_vendor_str(vendor_str):
    """解析厂家字符串"""
    vendors = []
    if not vendor_str.strip():
        return vendors
    
    for item in vendor_str.split("、"):
        item = item.strip()
        if "（" in item and "）" in item:
            name = item.split("（")[0].strip()
            gpu = item.split("（")[1].replace("）", "").strip()
            if name and gpu:
                vendors.append((name, gpu))
    
    return vendors


def set_cell_style(cell, styles, header=False):
    """设置单元格样式"""
    if header:
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
    cell.border = styles["border"]
    cell.alignment = styles["center_align"]


def auto_adjust_column_width(ws, max_width=50):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width