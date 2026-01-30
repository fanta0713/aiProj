# excel_export.py - Excel报告生成模块
import openpyxl
from datetime import datetime
from tkinter import filedialog, messagebox


class ExcelExporter:
    """负责生成Excel报告"""

    @staticmethod
    def generate_report(app_ref):
        """生成完整的Excel报告"""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"GPU性能测试-{datetime.now().strftime('%Y%m%d')}.xlsx",
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # 1. 项目信息
            ExcelExporter._write_project_info(wb, app_ref)
            # 2. 测试环境
            ExcelExporter._write_env_data(wb, app_ref)
            # 3. PK指标
            ExcelExporter._write_pk_data(wb, app_ref)
            # 4-6. 性能数据
            ExcelExporter._write_perf_data(wb, app_ref)
            # 7. 项目问题
            ExcelExporter._write_problem_data(wb, app_ref)
            # 8. 项目总结
            ExcelExporter._write_summary(wb, app_ref)

            wb.save(save_path)
            messagebox.showinfo("成功", f"Excel已生成：\n{save_path}")

        except Exception as e:
            messagebox.showerror("错误", f"生成失败：{str(e)}")

    @staticmethod
    def _write_project_info(wb, app_ref):
        """写入项目信息"""
        ws = wb.create_sheet("1. 项目信息", 0)
        ws["A1"] = "项目名称"
        ws["B1"] = app_ref.project_name.get()
        ws["A2"] = "测试周期"
        ws["B2"] = app_ref.test_cycle.get()
        ws["A3"] = "参与厂家"
        ws["B3"] = app_ref.vendor_str.get()
        ws["A4"] = "测试模型"
        ws["B4"] = "、".join(app_ref.selected_models)
        # 新增 Step1 中的项目字段
        ws["A6"] = "客户名称"
        ws["B6"] = app_ref.customer_name.get()
        ws["A7"] = "客户行业"
        ws["B7"] = app_ref.customer_industry.get()
        ws["A8"] = "中标情况"
        ws["B8"] = app_ref.bid_status.get()
        ws["A9"] = "中标份额"
        ws["B9"] = app_ref.bid_share.get()
        ws["A10"] = "未中标原因"
        ws["B10"] = app_ref.bid_fail_reason.get()
        ws["A11"] = "测试负责人"
        ws["B11"] = app_ref.test_owner.get()

    @staticmethod
    def _write_env_data(wb, app_ref):
        """写入测试环境数据"""
        ws = wb.create_sheet("2. 测试环境", 1)
        headers = ["序号", "模型", "测试类型", "厂家", "GPU配置", "GPU数量", "数据集", "测试工具"]
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)

        row = 2
        for idx, data in enumerate(app_ref.env_data):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, data["model"])
            ws.cell(row, 3, data["test_type"])
            ws.cell(row, 4, data["vendor"])
            ws.cell(row, 5, data["gpu"])
            ws.cell(row, 6, data["gpu_count"])
            ws.cell(row, 7, data["dataset"])
            ws.cell(row, 8, data["tool"])
            row += 1

    @staticmethod
    def _write_pk_data(wb, app_ref):
        """写入PK指标数据"""
        ws = wb.create_sheet("3. PK指标", 2)
        headers = ["序号", "模型", "测试类型", "PK指标"]
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)

        row = 2
        for idx, pk_row in enumerate(app_ref.pk_data):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, pk_row["model"])
            ws.cell(row, 3, pk_row["test_type"])
            ws.cell(row, 4, pk_row["selected_pk"])
            row += 1

    @staticmethod
    def _write_perf_data(wb, app_ref):
        """写入性能数据（推理、训练、精度）"""
        category_map = {
            "推理性能": [
                "文本推理",
                "图文推理",
                "图像识别",
                "语音推理",
                "文档排序",
                "特征提取",
            ],
            "训练性能": ["预训练", "lora微调", "全参微调"],
            "精度测试": ["精度测试"],
        }

        sheet_data = {
            "推理性能": {"sheet": wb.create_sheet("4. 推理性能数据", 3), "row": 1},
            "训练性能": {"sheet": wb.create_sheet("5. 训练性能数据", 4), "row": 1},
            "精度测试": {"sheet": wb.create_sheet("6. 精度测试数据", 5), "row": 1},
        }

        entry_idx = 0
        for perf_row in app_ref.perf_data:
            tt = perf_row["test_type"]
            category = None

            for cat, types in category_map.items():
                if tt in types:
                    category = cat
                    break

            if not category:
                continue

            info = sheet_data[category]
            ws = info["sheet"]
            row = info["row"]

            # 写表头（首次）
            if row == 1:
                base_headers = ["序号", "模型", "厂家", "数据集", "测试类型"]
                headers = base_headers + perf_row["input_fields"] + perf_row["calc_fields"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row, col, h)
                info["row"] += 1
                row += 1

            # 写数据行
            col = 1
            ws.cell(row, col, entry_idx)
            col += 1
            ws.cell(row, col, perf_row["model"])
            col += 1
            ws.cell(row, col, perf_row["vendor"])
            col += 1
            ws.cell(row, col, perf_row["dataset"])
            col += 1
            ws.cell(row, col, tt)
            col += 1

            for field in perf_row["input_fields"]:
                ws.cell(row, col, perf_row["input_values"][field])
                col += 1

            for field in perf_row["calc_fields"]:
                ws.cell(row, col, perf_row["calc_values"][field])
                col += 1

            info["row"] += 1
            entry_idx += 1

    @staticmethod
    def _write_problem_data(wb, app_ref):
        """写入项目问题数据"""
        ws = wb.create_sheet("7. 项目中遇到的问题", 6)
        problem_headers = ["序号", "问题分类", "问题描述", "责任人", "解决方案"]
        for col, h in enumerate(problem_headers, 1):
            ws.cell(1, col, h)

        row = 2
        for idx, problem_row in enumerate(app_ref.problem_data):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, problem_row["category"])
            ws.cell(row, 3, problem_row["description"])
            ws.cell(row, 4, problem_row["person"])
            ws.cell(row, 5, problem_row["solution"])
            row += 1

    @staticmethod
    def _write_summary(wb, app_ref):
        """写入项目总结"""
        ws = wb.create_sheet("8. 项目总结", 7)
        if app_ref.project_summary:
            summary_lines = app_ref.project_summary.split("\n")
            for row_idx, line in enumerate(summary_lines, 1):
                ws.cell(row_idx, 1, line)
        else:
            ws.cell(1, 1, "未生成项目总结，请先点击「生成项目总结」按钮生成")
